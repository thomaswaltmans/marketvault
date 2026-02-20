from django.shortcuts import render
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.conf import settings
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.db import IntegrityError
from django.db import transaction as db_transaction
from django.db.models import Q
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib.auth.decorators import login_required
from django.utils.dateparse import parse_datetime
from django.core.exceptions import ValidationError

import json
import logging
import re
from datetime import datetime, timezone as dt_timezone

from .models import User, Asset, Transaction

logger = logging.getLogger(__name__)

try:
    from django_ratelimit.decorators import ratelimit
except ModuleNotFoundError:
    # Local fallback if dependency is not installed yet.
    def ratelimit(*args, **kwargs):
        def decorator(func):
            return func
        return decorator

# Create your views here.
@ensure_csrf_cookie
def index(request):
    return render(request, "portfolio/index.html")


def _client_ip(request):
    forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "")

# Login views
@ratelimit(key="ip", rate="20/m", method="POST", block=False)
@ratelimit(key="post:username", rate="8/m", method="POST", block=False)
def login_view(request):
    if request.method == "POST":
        if getattr(request, "limited", False):
            logger.warning(
                "Rate limit hit on login_view ip=%s username=%s",
                _client_ip(request),
                request.POST.get("username", ""),
            )
            return render(request, "portfolio/login.html", {
                "message": "Too many login attempts. Please wait a minute and try again."
            }, status=429)

        # Attempt to sign user in
        username = request.POST["username"]
        password = request.POST["password"]
        user = authenticate(request, username=username, password=password)

        # Check if authentication successful
        if user is not None:
            login(request, user)
            return HttpResponseRedirect(reverse("index"))
        else:
            return render(request, "portfolio/login.html", {
                "message": "Invalid username and/or password."
            })
    else:
        return render(request, "portfolio/login.html")
    
def logout_view(request):
    logout(request)
    return HttpResponseRedirect(reverse("index"))

@ratelimit(key="ip", rate="10/m", method="POST", block=False)
def register(request):
    if not settings.REGISTRATION_ENABLED:
        return HttpResponse(status=404)

    if request.method == "POST":
        if getattr(request, "limited", False):
            logger.warning(
                "Rate limit hit on register ip=%s username=%s",
                _client_ip(request),
                request.POST.get("username", ""),
            )
            return render(request, "portfolio/register.html", {
                "message": "Too many signup attempts. Please wait and try again."
            }, status=429)

        username = request.POST["username"]
        email = request.POST["email"]

        # Ensure password matches confirmation
        password = request.POST["password"]
        confirmation = request.POST["confirmation"]
        if password != confirmation:
            return render(request, "portfolio/register.html", {
                "message": "Passwords must match."
            })

        # Attempt to create new user
        try:
            user = User.objects.create_user(username, email, password)
            user.save()
        except IntegrityError:
            return render(request, "portfolio/register.html", {
                "message": "Username already taken."
            })
        login(request, user)
        return HttpResponseRedirect(reverse("index"))
    else:
        return render(request, "portfolio/register.html")
    
# Transaction views
@login_required
def transactions(request):
    if request.method == "GET":
        transactions = (Transaction.objects.filter(user=request.user).select_related("asset").order_by("-timestamp"))
        return JsonResponse({"transactions": [transaction.serialize() for transaction in transactions]})

    if request.method == "POST":
        data = json.loads(request.body or "{}")

        transaction_type = data.get("txn_type")
        asset_id = data.get("asset_id")

        if not transaction_type:
            return JsonResponse({"error": "txn_type is required"}, status=400)
        if not asset_id:
            return JsonResponse({"error": "asset_id is required"}, status=400)

        try:
            asset = Asset.objects.get(id=asset_id, user=request.user)
        except Asset.DoesNotExist:
            return JsonResponse({"error": "Asset not found"}, status=404)

        transaction = Transaction(
            user=request.user,
            txn_type=transaction_type,
            asset=asset,
            quantity=data.get("quantity"),
            unit_price=data.get("unit_price"),
            div_amount=data.get("div_amount"),
        )

        if "timestamp" in data:
            transaction.timestamp = parse_datetime(data["timestamp"])

        try:
            transaction.save()
        except ValidationError as error:
            return JsonResponse({"errors": error.message_dict}, status=400)

        return JsonResponse(transaction.serialize(), status=201)

@login_required
def transaction(request, transaction_id):
    try:
        transaction = Transaction.objects.select_related("asset").get(id=transaction_id, user=request.user)
    except Transaction.DoesNotExist:
        return JsonResponse({"error": "Transaction not found"}, status=404)

    if request.method == "GET":
        return JsonResponse(transaction.serialize())

    if request.method == "PUT":
        data = json.loads(request.body or "{}")

        for field in ["txn_type", "quantity", "unit_price", "div_amount"]:
            if field in data:
                setattr(transaction, field, data[field])

        if "asset_id" in data:
            try:
                transaction.asset = Asset.objects.get(id=data["asset_id"], user=request.user)
            except Asset.DoesNotExist:
                return JsonResponse({"error": "Asset not found"}, status=404)

        if "timestamp" in data:
            parsed_timestamp = parse_datetime(data["timestamp"])
            if not parsed_timestamp:
                return JsonResponse({"error": "Invalid timestamp"}, status=400)
            transaction.timestamp = parsed_timestamp

        try:
            transaction.save()
        except ValidationError as error:
            return JsonResponse({"errors": error.message_dict}, status=400)

        return JsonResponse(transaction.serialize())

    if request.method == "DELETE":
        transaction.delete()
        return JsonResponse({"message": "Deleted"})

    return JsonResponse({"error": "GET, PUT, DELETE required"}, status=405)


@login_required
def assets(request):
    # GET: list / search assets
    if request.method == "GET":
        query = request.GET.get("q", "").strip()

        assets = Asset.objects.filter(user=request.user).order_by("ticker")

        if query:
            assets = assets.filter(
                Q(ticker__icontains=query) |
                Q(name__icontains=query) |
                Q(data_symbol__icontains=query)
            )

        asset_list = []
        for asset in assets[:50]:
            asset_list.append({
                "id": asset.id,
                "ticker": asset.ticker,
                "name": asset.name,
                "asset_type": asset.asset_type,
                "exchange": asset.exchange,
                "currency": asset.currency,
                "data_symbol": asset.data_symbol,
            })

        return JsonResponse({"assets": asset_list})

    # POST: create asset
    if request.method == "POST":
        try:
            data = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        ticker = (data.get("ticker") or "").strip().upper()
        name = (data.get("name") or "").strip()
        asset_type = (data.get("asset_type") or Asset.AssetType.STOCK).strip().upper()
        exchange = (data.get("exchange") or "").strip()
        currency = (data.get("currency") or "EUR").strip().upper()
        data_symbol = (data.get("data_symbol") or "").strip()
        valid_types = set(Asset.AssetType.values)

        if not ticker:
            return JsonResponse({"error": "ticker is required"}, status=400)

        if not data_symbol:
            return JsonResponse({"error": "data_symbol is required"}, status=400)
        
        if asset_type not in valid_types:
            return JsonResponse({"error": "asset_type must be one of ETF, STOCK, ETC, CRYPTO"}, status=400)

        try:
            asset = Asset.objects.create(
                user=request.user,
                ticker=ticker,
                name=name,
                asset_type=asset_type,
                exchange=exchange,
                currency=currency,
                data_symbol=data_symbol,
            )
        except ValidationError as error:
            return JsonResponse({"errors": error.message_dict}, status=400)
        except IntegrityError:
            return JsonResponse({"error": "Asset already exists for this account"}, status=400)

        return JsonResponse({
            "id": asset.id,
            "ticker": asset.ticker,
            "name": asset.name,
            "asset_type": asset.asset_type,
            "exchange": asset.exchange,
            "currency": asset.currency,
            "data_symbol": asset.data_symbol,
        }, status=201)

    return JsonResponse({"error": "GET or POST required"}, status=405)

@login_required
def asset(request, asset_id):
    try:
        asset = Asset.objects.get(id=asset_id, user=request.user)
    except Asset.DoesNotExist:
        return JsonResponse({"error": "Asset not found"}, status=404)

    # GET: detail
    if request.method == "GET":
        return JsonResponse({
            "id": asset.id,
            "ticker": asset.ticker,
            "name": asset.name,
            "asset_type": asset.asset_type,
            "exchange": asset.exchange,
            "currency": asset.currency,
            "data_symbol": asset.data_symbol,
        })

    # PUT: update
    if request.method == "PUT":
        try:
            data = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        if "ticker" in data:
            asset.ticker = data["ticker"].strip().upper()

        if "name" in data:
            asset.name = data["name"].strip()
        
        if "asset_type" in data:
            next_type = data["asset_type"].strip().upper()
            if next_type not in set(Asset.AssetType.values):
                return JsonResponse({"error": "asset_type must be one of ETF, STOCK, ETC, CRYPTO"}, status=400)
            asset.asset_type = next_type

        if "exchange" in data:
            asset.exchange = data["exchange"].strip()

        if "currency" in data:
            asset.currency = data["currency"].strip().upper()

        if "data_symbol" in data:
            asset.data_symbol = data["data_symbol"].strip()

        try:
            asset.full_clean()
            asset.save()
        except ValidationError as error:
            return JsonResponse({"errors": error.message_dict}, status=400)
        except IntegrityError:
            return JsonResponse({"error": "Asset already exists for this account"}, status=400)

        return JsonResponse({
            "id": asset.id,
            "ticker": asset.ticker,
            "name": asset.name,
            "asset_type": asset.asset_type,
            "exchange": asset.exchange,
            "currency": asset.currency,
            "data_symbol": asset.data_symbol,
        })

    # DELETE
    if request.method == "DELETE":
        if Transaction.objects.filter(asset=asset).exists():
            return JsonResponse(
                {"error": "Cannot delete asset with existing transactions"},
                status=400
            )

        asset.delete()
        return JsonResponse({"message": "Deleted"})

    return JsonResponse({"error": "GET, PUT, or DELETE required"}, status=405)


@login_required
def profile(request):
    user = request.user

    if request.method == "GET":
        return JsonResponse({
            "username": user.username,
            "email": user.email,
            "first_name": user.first_name,
            "last_name": user.last_name,
        })

    if request.method == "PUT":
        try:
            data = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON"}, status=400)

        if "email" in data:
            user.email = data["email"].strip()

        if "first_name" in data:
            user.first_name = data["first_name"].strip()

        if "last_name" in data:
            user.last_name = data["last_name"].strip()

        try:
            user.full_clean()
            user.save()
        except ValidationError as error:
            return JsonResponse({"errors": error.message_dict}, status=400)

        return JsonResponse({
            "username": user.username,
            "email": user.email,
            "first_name": user.first_name,
            "last_name": user.last_name,
        })

    return JsonResponse({"error": "GET or PUT required"}, status=405)


@ratelimit(key="user_or_ip", rate="10/m", method="POST", block=False)
@login_required
def profile_password(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    if getattr(request, "limited", False):
        logger.warning(
            "Rate limit hit on profile_password ip=%s user_id=%s username=%s",
            _client_ip(request),
            getattr(request.user, "id", None),
            getattr(request.user, "username", ""),
        )
        return JsonResponse(
            {"error": "Too many password change attempts. Please wait and try again."},
            status=429,
        )

    try:
        data = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    current_password = data.get("current_password") or ""
    new_password = data.get("new_password") or ""
    confirm_password = data.get("confirm_password") or ""

    if not request.user.check_password(current_password):
        return JsonResponse({"error": "Current password is incorrect"}, status=400)

    if len(new_password) < 8:
        return JsonResponse({"error": "New password must be at least 8 characters"}, status=400)

    if new_password != confirm_password:
        return JsonResponse({"error": "New password and confirmation do not match"}, status=400)

    request.user.set_password(new_password)
    request.user.save()
    update_session_auth_hash(request, request.user)

    return JsonResponse({"message": "Password updated"})

### IMPORTING FROM CSV AND HELPER FUNCTIONS

def _normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", "_", text)  # all whitespace -> underscore
    return text


def _clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _clean_decimal(value):
    """
    Takes Excel values like '€415.10' or '415,10' and returns a string like '415.10'.
    Returns None if empty.
    """
    if value is None or value == "":
        return None

    text = str(value).strip()
    text = text.replace("€", "").replace("$", "").replace("£", "")
    text = text.replace(" ", "")

    # comma decimal (EU) -> dot
    if "," in text and "." not in text:
        text = text.replace(",", ".")

    # thousand separators
    if "," in text and "." in text:
        text = text.replace(",", "")

    return text or None


def _parse_unix_timestamp(value):
    if value is None or str(value).strip() == "":
        return None

    text = str(value).strip()
    if not text.isdigit():
        return None

    seconds = int(text)
    dt_utc = datetime.fromtimestamp(seconds, tz=dt_timezone.utc)
    return dt_utc.astimezone(timezone.get_current_timezone())


def _derive_ticker(data_symbol):
    return data_symbol.split(".")[0].strip().upper()


@login_required
def import_data(request):
    if request.method != "POST":
        return JsonResponse({"error": "POST required"}, status=405)

    if "file" not in request.FILES:
        return JsonResponse({"error": "No file uploaded (field name should be 'file')"}, status=400)

    uploaded_file = request.FILES["file"]
    if not (uploaded_file.name or "").lower().endswith(".xlsx"):
        return JsonResponse({"error": "Please upload an .xlsx file"}, status=400)

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return JsonResponse(
            {"error": "Excel import is unavailable because openpyxl is not installed"},
            status=500,
        )

    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = workbook.active

    header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [_normalize_header(h) for h in header_row]
    header_set = set([h for h in headers if h])

    required = {"data_symbol", "txn_type", "timestamp"}
    missing = required - header_set
    if missing:
        return JsonResponse({"error": f"Missing required columns: {sorted(list(missing))}"}, status=400)

    created_assets = 0
    created_transactions = 0
    row_errors = []

    for excel_row_index, row_values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = dict(zip(headers, row_values))

        data_symbol = _clean_text(row.get("data_symbol"))
        txn_type = _clean_text(row.get("txn_type")).upper()
        ts = _parse_unix_timestamp(row.get("timestamp"))

        quantity = _clean_decimal(row.get("quantity"))
        unit_price = _clean_decimal(row.get("unit_price"))
        div_amount = _clean_decimal(row.get("div_amount"))

        if not data_symbol or not txn_type:
            row_errors.append({"row": excel_row_index, "error": "data_symbol and txn_type are required"})
            continue

        if ts is None:
            row_errors.append({"row": excel_row_index, "error": "timestamp must be unix seconds (e.g. 1610323200)"})
            continue

        asset, was_created = Asset.objects.get_or_create(
            user=request.user,
            data_symbol=data_symbol,
            defaults={
                "ticker": _derive_ticker(data_symbol),
                "name": "",
                "exchange": "",
                "currency": "EUR",
                "asset_type": Asset.AssetType.STOCK,
            }
        )
        if was_created:
            created_assets += 1

        txn = Transaction(
            user=request.user,
            asset=asset,
            txn_type=txn_type,
            quantity=quantity,
            unit_price=unit_price,
            div_amount=div_amount,
            timestamp=ts,
        )

        try:
            txn.save()  # model save calls full_clean()
            created_transactions += 1
        except Exception as e:
            row_errors.append({"row": excel_row_index, "error": str(e)})

    return JsonResponse({
        "created_transactions": created_transactions,
        "created_assets": created_assets,
        "row_errors": row_errors,
    }, status=201)


### ANALYTICS

@login_required
def analytics_growth(request):
    if request.method != "GET":
        return JsonResponse({"error": "GET required"}, status=405)

    try:
        from portfolio.services.analytics import growth_payload
    except ModuleNotFoundError:
        return JsonResponse({"error": "Analytics is unavailable because pandas is not installed"}, status=500)

    payload = growth_payload(request.user)
    return JsonResponse(payload)


@login_required
def analytics_allocation(request):
    if request.method != "GET":
        return JsonResponse({"error": "GET required"}, status=405)

    try:
        from portfolio.services.analytics import allocation_payload
    except ModuleNotFoundError:
        return JsonResponse({"error": "Analytics is unavailable because pandas is not installed"}, status=500)

    payload = allocation_payload(request.user)
    return JsonResponse(payload)


@login_required
def analytics_asset_growth(request):
    if request.method != "GET":
        return JsonResponse({"error": "GET required"}, status=405)

    try:
        from portfolio.services.analytics import asset_growth_payload
    except ModuleNotFoundError:
        return JsonResponse({"error": "Analytics is unavailable because pandas is not installed"}, status=500)

    payload = asset_growth_payload(request.user)
    return JsonResponse(payload)
